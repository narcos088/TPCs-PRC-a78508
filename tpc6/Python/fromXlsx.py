#!/usr/bin/python3

import openpyxl,json

jsonO = {"d":[]}
doc = openpyxl.load_workbook('freguesias-metadata.xlsx')

sheet = doc.active
max_row = sheet.max_row
max_column = sheet.max_column

for r in range(2,max_row+1):
    entry ={}
    entry['PartitionKey']='Sheet2'
    for c in range(1,max_column+1):
        key = sheet.cell(row=1,column=c).value 
        value = sheet.cell(row=r,column=c).value
        #print(str(key)+":"+str(value))
        #print()
        if(value):
            entry[key] = value
    jsonO['d'].append(entry)

with open('freguesias.json', 'w') as outfile:
    json.dump(jsonO, outfile,indent=4,ensure_ascii=False)
